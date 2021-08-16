# coding=UTF-8
import datetime
import json
import os
from functools import reduce

import openpyxl
from chinese_calendar import is_workday
from redminelib import Redmine

# 查找对应的元组数据
find_index = lambda self, i, value: sorted(self, key=lambda x: x[i] != value)[0]

# pm host
pm_host = 'https://pm.nexttao.com'


def save_file(overtimes):
    with open('./data.json', 'w') as f:
        f.write(json.dumps(overtimes))


def format_emit(arr):
    # 存入 文件
    save_file(arr)

    count_overtime = reduce(
        lambda x, y: x + y,
        list(map(lambda x: x['overtime'], arr))
    )

    res_json = json.dumps({
        "items": [
            {
                'title': f'当月加班 {count_overtime} 小时',
                'subtitle': '回车导出为 excel',
                'arg': '0',  # 必须的参数，只有有这个参数，才可以链接下一个 workflow 动作
                'valid': True
            }
        ]
    })

    print(res_json)


def export():
    with open('./data.json', 'r') as f:
        data = f.read()

        last_computed_overtimes = json.loads(data)

    # 创建一个 excel
    today = datetime.datetime.today()
    year = today.year
    month = today.month

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '工时统计'

    sheet.page_setup.fitToWidth = 0

    # 第一行
    row_title = [
        ('username', '用户名'),
        ('spent_on', '日期'),
        ('overtime', '加班时长'),
        ('remarks', '备注'),
        ('project', '项目')
    ]
    for index, column_name in enumerate(row_title):
        sheet.cell(row=1, column=index + 1, value=column_name[1])

    for i in range(0, len(last_computed_overtimes)):
        item = last_computed_overtimes[i]

        for index, column_name in enumerate(row_title):
            sheet.cell(row=i + 2, column=index + 1, value=str(item[column_name[0]]))

    home = os.path.expanduser('~')

    path = f'{home}/Downloads/工时统计-{year}-{month}.xlsx'

    workbook.save(path)

    # print(last_computed_overtimes)


class RedmineQuery:
    # 初始化数据
    def __init__(self, key):
        # 初始化 redmine 资源
        if key == '' or key is None:
            key = os.environ.get('key')
            
        self.redmine = Redmine(pm_host, key)

        self.current_user = dict(self.redmine.user.get('current'))

        self.id = self.current_user['id']

    # 统计当月工时
    def get_person_hour(self):
        # print(redmine)
        # 上个月开始的第一天
        last_month_start = datetime.date(datetime.date.today().year, datetime.date.today().month - 1, 1)

        current_month_start = datetime.date(datetime.date.today().year, datetime.date.today().month, 1)

        time_entries = self.redmine.time_entry.filter(
            user_id=self.id,
            from_date=current_month_start,
            limit=1000
        )

        time_items = map(lambda x: list(x), list(time_entries))

        # 按发生日期分组
        items_by_group = {}

        # print(list(time_items))

        for time_item in time_items:
            spent_on = find_index(time_item, 0, 'spent_on')[1]
            username = find_index(time_item, 0, 'user')[1]['name']
            project = find_index(time_item, 0, 'project')[1]['name']
            hours = find_index(time_item, 0, 'hours')[1]
            issue = find_index(time_item, 0, 'issue')[1]['id']
            comments = find_index(time_item, 0, 'comments')[1]

            new_comments = f' ({comments})' if comments != '' else ''

            remarks = f'{pm_host}/issues/{issue}{new_comments}'

            # 如果日期存在， 则时间相加
            if spent_on in items_by_group:

                # 如果当前记录用时偏大，则优先使用当前记录的备注和项目
                if hours > items_by_group[spent_on]['hours']:
                    items_by_group[spent_on]['remarks'] = remarks
                    items_by_group[spent_on]['project'] = project

                # 时间相加
                items_by_group[spent_on]['hours'] = items_by_group[spent_on]['hours'] + hours

            else:
                # 如果日期不存在，直接添加
                items_by_group[spent_on] = {
                    'username': username,
                    'spent_on': spent_on,
                    'project': project,
                    'hours': hours,
                    'remarks': remarks
                }

            # 计算加班时间

            # 如果是非工作日, 直接计算为加班时间
            if not is_workday(datetime.datetime.strptime(spent_on, '%Y-%m-%d')):
                items_by_group[spent_on]['overtime'] = items_by_group[spent_on]['hours']
            else:
                # 如果工作日，则计算是否超过 8h
                if items_by_group[spent_on]['hours'] > 8:
                    items_by_group[spent_on]['overtime'] = items_by_group[spent_on]['hours'] - 8

        # 过滤出加班数据
        overtimes = list(filter(lambda x: 'overtime' in x.keys(), items_by_group.values()))

        format_emit(overtimes)

    # 查询 issue info
    def get_issue_info(self, issue_id):
        issues = dict(self.redmine.issue.get(issue_id))

        res_json = json.dumps({
            "items": [
                {
                    'title': f'{issues["subject"]}',
                    'subtitle': f'{issues["description"]} 回车查看详情',
                    'arg': issue_id
                }
            ]
        })

        print(res_json)

    pass
